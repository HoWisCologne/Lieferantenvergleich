#!/usr/bin/env python3
"""
Liest die Excel-Preisliste ein und erzeugt den Seed-Datensatz fuer die
Lieferantenvergleich-Anwendung (seed.js + seed.json).

Datenmodell:
  - lieferanten : Firmen (Buender, Fassbender, Zaun, Henrich)
  - artikel     : logischer Artikel (Name + Laenge + Nennweite + Grad + Gewicht)
                  -> identisch ueber alle Lieferanten, damit verglichen werden kann
  - angebote    : Verknuepfung Lieferant<->Artikel inkl. lieferantenspezifischer
                  Artikelnummer und Einheit
  - preise      : Preishistorie je Angebot (preis, datum, notiz)

Erneut ausfuehren mit:  python3 build_seed.py
"""
import openpyxl, json, re, datetime, os, hashlib

XLSX = "/Users/pc/Library/CloudStorage/OneDrive-HorstWiskirchenFuhr-undBaggerbetrieb/Firma Horst Wiskirchen/versch. Firmen/a Preislisten/a Listen für Bünder,Fassbender und Zaun/Preisliste Bünder, Zaun & Fassbender Tenten.xlsx"
HERE = os.path.dirname(os.path.abspath(__file__))


def clean(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return re.sub(r"\s+", " ", v).strip()
    return v


def to_iso(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return None


# --- Kategorisierung in grobe Überbegriffe (Reihenfolge = Priorität) ----------
# Identische Logik liegt auch in index.html (kategorieFor), damit vorhandene
# Daten und neu angelegte Artikel gleich einsortiert werden.
KAT_REGELN = [
    ("Verpackung, Fracht & Service", ["palette", "gitterbox", "vorfracht", "fracht",
        "anfuhr", "maut", "lieferservice", "kranservice", "express", "stückgut",
        "stueckgut", "palettengeb"]),
    ("Erdung & Blitzschutz", ["erder", "erdungsband", "ringerder", "runddraht"]),
    ("Kabelschutz", ["kabuflex", "kabel"]),
    ("Abdichtung, Folien & Noppenbahnen", ["folie", "baufolie", "noppenbahn", "delta",
        "nevobahn", "nevo", "terraxx", "terrraxx", "bitumen", "dickbeschicht",
        "dicktbeschicht"]),
    ("Vliese & Geogitter", ["vlies", "geotextil", "geotext", "geovlies", "geogitter",
        "polyfelt", "betex", "tensar", "begrid", "galatex", "evalith", "wurzelvlies",
        "unkrautvlies", "ag 2 rolle", "ag 5 rollen"]),
    ("Drainage & Dränage", ["opti-control", "opti.control", "opti control", "opti-drän",
        "opti-dran", "opti drän", "drän", "dräna", "draen", "drainrohr", "drainagerohr",
        "drain-ff", "kokofill", "rngff"]),
    ("Schächte & Schachtbauteile", ["schacht", "sx400", "sx 400", "awaschacht",
        "auflagering", "betondeckel", "aufsetzrohr", "asb", "aufsatz", "lichtschacht",
        "teleskop", "konus", "steigeisen", "sicherheitseisen"]),
    ("Bordsteine, Pflaster & Einfassungen", ["tiefboard", "tiefbord", "bordstein",
        "randstein", "rundboard", "hochboard", "beeteinfass", "basaltin", "pflaster",
        "fugenband", "board"]),
    ("Beton, Zement & Mörtel", ["zement", "estrich", "mörtel", "moertel", "mortel",
        "trockenbeton", "schnellbeton", "schnellzement", "trasszement", "baumit",
        "botament", "remix", "weber.rep", "betonstahl", "betonmischung",
        "kalksandstein", "beton"]),
    ("Zubehör & Hilfsmittel", ["gleitmittel", "handschuh", "plattennäg", "plattennag",
        "soudal", "brunnenschaum", "kraso", "schlauchschell", "schlauchverbind",
        "reduzierring"]),
    ("Kanalrohre & Formstücke", ["kg2000", "kg 2000", "kg-", "kgus", "kgb", "kgm",
        "kgr", "ht ", "htr", "htb", "htea", "ht-", "rohr", "bogen", "abzweig", "muffe",
        "spitze", "reduz", "redu", "überschieb", "ueberschieb", "verschlußdeckel",
        "verschlussdeckel", "veschlußdeckel", "reinigungsrohr", "rückstau", "rueckstau",
        "ruckstau", "stopfen", "steinzeug", "tonmuffe", "tonsp", "kanalrohr", "kanalr",
        "pvc", "aco"]),
]


def kategorie_for(name):
    n = " " + name.lower() + " "
    for cat, keys in KAT_REGELN:
        for k in keys:
            if k in n:
                return cat
    return "Sonstiges"


def guess_unit(name):
    n = name.lower()
    if any(k in n for k in ["folie", "vlies", "bahn", "geogitter", "geo-drain",
                            "geo drain", "geotextil", "noppenbahn", "wurzelvlies",
                            "unkrautvlies", "terraxx", "polyfelt", "kabuflex",
                            "kokofill", "drainrohr", "draenrohr"]):
        return "Rolle"
    if any(k in n for k in ["zement", "estrich", "moertel", "mörtel", "trockenbeton",
                            "schnellbeton", "schnellzement", "botament", "putz",
                            "körnung", "koernung", "betonmischung"]):
        return "Sack"
    if "eimer" in n:
        return "Eimer"
    if "tube" in n or "gleitmittel" in n:
        return "Tube"
    if "palette" in n or "gitterbox" in n:
        return "Palette"
    return "Stk."


# ---- ID-Helfer: stabile IDs auf Basis des Inhalts ---------------------------
def aid(name, laenge, nennweite, grad, gewicht):
    key = "|".join(str(x) for x in [name.lower(), laenge, nennweite, grad, gewicht])
    return "a_" + hashlib.md5(key.encode("utf-8")).hexdigest()[:10]


wb = openpyxl.load_workbook(XLSX, data_only=True)

lieferanten = [
    {"id": "l_buender",    "name": "Bünder",     "notiz": ""},
    {"id": "l_fassbender", "name": "Fassbender",  "notiz": ""},
    {"id": "l_zaun",       "name": "Zaun",        "notiz": ""},
    {"id": "l_henrich",    "name": "Henrich",     "notiz": "Preise inkl. 76 % Nachlass (lt. Vergleichsliste)"},
]

artikel = {}        # aid -> artikel dict
angebote = []       # angebot dicts
preise = []         # preis dicts
ang_counter = 0
preis_counter = 0


def add_entry(lief_id, name, artnr, laenge, nennweite, grad, gewicht,
              preis, stand_iso, altpreis):
    """Legt logischen Artikel (falls neu), Angebot und Preishistorie an."""
    global ang_counter, preis_counter
    a = aid(name, laenge, nennweite, grad, gewicht)
    if a not in artikel:
        artikel[a] = {
            "id": a, "name": name, "laenge": laenge, "nennweite": nennweite,
            "grad": grad, "gewicht": gewicht, "kategorie": kategorie_for(name),
        }
    ang_counter += 1
    ang_id = f"ang_{ang_counter:04d}"
    angebote.append({
        "id": ang_id, "lieferantId": lief_id, "artikelId": a,
        "artikelnummer": str(artnr) if artnr not in (None, "") else "",
        "einheit": guess_unit(name), "notiz": "",
    })
    # Preishistorie: ggf. Altpreis als aelterer Eintrag, dann aktueller Preis
    if isinstance(altpreis, (int, float)) and abs(float(altpreis) - float(preis)) > 1e-9:
        try:
            d = datetime.datetime.strptime(stand_iso, "%Y-%m-%d") if stand_iso else None
        except Exception:
            d = None
        alt_iso = (d - datetime.timedelta(days=180)).strftime("%Y-%m-%d") if d else None
        preis_counter += 1
        preise.append({
            "id": f"p_{preis_counter:05d}", "angebotId": ang_id,
            "preis": round(float(altpreis), 2), "datum": alt_iso,
            "notiz": "Altpreis (Import aus Liste)",
        })
    preis_counter += 1
    preise.append({
        "id": f"p_{preis_counter:05d}", "angebotId": ang_id,
        "preis": round(float(preis), 2), "datum": stand_iso,
        "notiz": "Import aus Preisliste",
    })


# ---- Lieferanten-Sheets (Buender / Fassbender / Zaun) -----------------------
sheet_map = {"Bünder": "l_buender", "Fassbender": "l_fassbender", "Zaun": "l_zaun"}
for sheet, lief_id in sheet_map.items():
    ws = wb[sheet]
    for r in range(3, ws.max_row + 1):
        name = clean(ws.cell(r, 1).value)
        preis = ws.cell(r, 7).value
        if not name or not isinstance(preis, (int, float)):
            continue
        artnr = clean(ws.cell(r, 2).value)
        laenge = clean(ws.cell(r, 3).value)
        nennweite = clean(ws.cell(r, 4).value)
        grad_raw = clean(ws.cell(r, 5).value)
        grad = grad_raw if isinstance(grad_raw, str) and "grad" in grad_raw.lower() else (
            grad_raw if isinstance(grad_raw, (int, float)) else "")
        gew_raw = ws.cell(r, 6).value
        gewicht = gew_raw if isinstance(gew_raw, (int, float)) else ""
        stand = to_iso(ws.cell(r, 8).value)
        altpreis = ws.cell(r, 9).value
        add_entry(lief_id, name, artnr, str(laenge), str(nennweite),
                  str(grad), gewicht, float(preis), stand, altpreis)

# ---- Henrich aus dem Vergleich-Sheet (Spalte L = Preis, N = Stand) ----------
ws = wb["Vergleich"]
for r in range(3, ws.max_row + 1):
    name = clean(ws.cell(r, 1).value)
    preis = ws.cell(r, 12).value  # L
    if not name or not isinstance(preis, (int, float)):
        continue
    laenge = clean(ws.cell(r, 2).value)      # B
    nennweite = clean(ws.cell(r, 3).value)   # C
    grad_raw = clean(ws.cell(r, 4).value)    # D
    grad = grad_raw if isinstance(grad_raw, str) and "grad" in grad_raw.lower() else ""
    gew_raw = ws.cell(r, 5).value            # E
    gewicht = gew_raw if isinstance(gew_raw, (int, float)) else ""
    stand = to_iso(ws.cell(r, 14).value)     # N
    add_entry("l_henrich", name, "", str(laenge), str(nennweite),
              str(grad), gewicht, float(preis), stand, None)

einheiten = ["Stk.", "Rolle", "lfm", "m", "m²", "Sack", "Eimer", "Tube",
             "Palette", "kg", "Karton", "Pauschal"]

data = {
    "version": 1,
    "generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
    "einheiten": einheiten,
    "lieferanten": lieferanten,
    "artikel": list(artikel.values()),
    "angebote": angebote,
    "preise": preise,
}

with open(os.path.join(HERE, "seed.json"), "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=1)

with open(os.path.join(HERE, "seed.js"), "w", encoding="utf-8") as f:
    f.write("// Automatisch generiert von build_seed.py – nicht von Hand bearbeiten.\n")
    f.write("window.SEED = ")
    json.dump(data, f, ensure_ascii=False)
    f.write(";\n")

print("Lieferanten :", len(lieferanten))
print("Artikel     :", len(artikel))
print("Angebote    :", len(angebote))
print("Preise      :", len(preise))
print("-> seed.json & seed.js geschrieben")
